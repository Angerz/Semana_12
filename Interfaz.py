# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'eso_elemetos.ui'
#
# Created by: PyQt5 UI code generator 5.9.2
#
# WARNING! All changes made in this file will be lost!

from PyQt5 import QtCore, QtGui, QtWidgets
import numpy as np
from tablas_fajas import Tabla_17_13, Tabla_17_12_C, Tabla_17_15, Tabla_17_10, Tabla17_14
from openpyxl.styles import Alignment, Border, Side
from openpyxl import load_workbook,Workbook
import os
from interfaz_tablas import ImageGallery


class Ui_MainWindow_principal(object):
    def setupUi(self, MainWindow_principal):
        MainWindow_principal.setObjectName("MainWindow_principal")
        MainWindow_principal.resize(731, 580)
        self.centralwidget = QtWidgets.QWidget(MainWindow_principal)
        self.centralwidget.setObjectName("centralwidget")
        self.frame = QtWidgets.QFrame(self.centralwidget)
        self.frame.setGeometry(QtCore.QRect(0, 0, 221, 579))
        self.frame.setStyleSheet("background-color: rgb(118, 118, 118);")
        self.frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame.setObjectName("frame")
        self.Button_Ingresa = QtWidgets.QPushButton(self.frame)
        self.Button_Ingresa.setGeometry(QtCore.QRect(0, 90, 221, 51))
        self.Button_Ingresa.setStyleSheet("QPushButton{\n"
"background-color: rgb(255, 255, 255);\n"
"color: rgb(0, 0, 0);\n"
"font: 10pt \"Times New Roman\";\n"
"}\n"
"\n"
"QPushButton:hover{\n"
"background-color: rgb(0,0,0);\n"
"color: rgb(255, 255, 255);\n"
"font: 10pt \"Times New Roman\";\n"
"}")
        self.Button_Ingresa.setObjectName("Button_Ingresa")
        self.Button_Resultado = QtWidgets.QPushButton(self.frame)
        self.Button_Resultado.setGeometry(QtCore.QRect(0, 210, 221, 51))
        self.Button_Resultado.setStyleSheet("QPushButton{\n"
"background-color: rgb(255, 255, 255);\n"
"color: rgb(0, 0, 0);\n"
"font: 10pt \"Times New Roman\";\n"
"}\n"
"\n"
"QPushButton:hover{\n"
"background-color: rgb(0,0,0);\n"
"color: rgb(255, 255, 255);\n"
"font: 10pt \"Times New Roman\";\n"
"}")
        self.Button_Resultado.setObjectName("Button_Resultado")
        self.Button_Tip_maq = QtWidgets.QPushButton(self.frame)
        self.Button_Tip_maq.setGeometry(QtCore.QRect(0, 330, 221, 51))
        self.Button_Tip_maq.setStyleSheet("QPushButton{\n"
"background-color: rgb(255, 255, 255);\n"
"color: rgb(0, 0, 0);\n"
"font: 10pt \"Times New Roman\";\n"
"}\n"
"\n"
"QPushButton:hover{\n"
"background-color: rgb(86, 179, 240);\n"
"color: rgb(255, 255, 255);\n"
"font: 10pt \"Times New Roman\";\n"
"}")
        self.Button_Tip_maq.setObjectName("Button_Tip_maq")
        self.Button_TAB_RES = QtWidgets.QPushButton(self.frame)
        self.Button_TAB_RES.setGeometry(QtCore.QRect(0, 430, 221, 51))
        self.Button_TAB_RES.setStyleSheet("QPushButton{\n"
"background-color: rgb(255, 255, 255);\n"
"color: rgb(0, 0, 0);\n"
"font: 10pt \"Times New Roman\";\n"
"}\n"
"\n"
"QPushButton:hover{\n"
"background-color: rgb(86, 179, 240);\n"
"color: rgb(255, 255, 255);\n"
"font: 10pt \"Times New Roman\";\n"
"}")
        self.Button_TAB_RES.setObjectName("Button_TAB_RES")
        self.Button_Resultado.raise_()
        self.Button_Tip_maq.raise_()
        self.Button_Ingresa.raise_()
        self.Button_TAB_RES.raise_()
        self.frame_5 = QtWidgets.QFrame(self.centralwidget)
        self.frame_5.setGeometry(QtCore.QRect(720, 760, 511, 281))
        self.frame_5.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_5.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_5.setObjectName("frame_5")
        self.verticalLayout = QtWidgets.QVBoxLayout(self.frame_5)
        self.verticalLayout.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout.setSpacing(0)
        self.verticalLayout.setObjectName("verticalLayout")
        self.frame_4 = QtWidgets.QFrame(self.centralwidget)
        self.frame_4.setGeometry(QtCore.QRect(220, 140, 509, 279))
        self.frame_4.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.frame_4.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_4.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_4.setObjectName("frame_4")
        self.label_18 = QtWidgets.QLabel(self.frame_4)
        self.label_18.setGeometry(QtCore.QRect(50, 30, 411, 51))
        self.label_18.setStyleSheet("font: 24pt \"Times New Roman\";")
        self.label_18.setObjectName("label_18")
        self.tableWidget = QtWidgets.QTableWidget(self.frame_4)
        self.tableWidget.setGeometry(QtCore.QRect(0, 90, 509, 279))
        self.tableWidget.setTabletTracking(False)
        self.tableWidget.setContextMenuPolicy(QtCore.Qt.NoContextMenu)
        self.tableWidget.setStyleSheet("")
        self.tableWidget.setAlternatingRowColors(False)
        self.tableWidget.setObjectName("tableWidget")
        self.tableWidget.setColumnCount(2)
        self.tableWidget.setRowCount(4)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setVerticalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setVerticalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setVerticalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setVerticalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        item.setFlags(QtCore.Qt.ItemIsSelectable|QtCore.Qt.ItemIsEditable|QtCore.Qt.ItemIsDragEnabled|QtCore.Qt.ItemIsUserCheckable|QtCore.Qt.ItemIsEnabled)
        self.tableWidget.setItem(0, 0, item)
        item = QtWidgets.QTableWidgetItem()
        item.setText("1.1 a 1.3")
        self.tableWidget.setItem(0, 1, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setItem(1, 0, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setItem(1, 1, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setItem(2, 0, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setItem(2, 1, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setItem(3, 0, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setItem(3, 1, item)
        self.tableWidget.horizontalHeader().setVisible(True)
        self.tableWidget.horizontalHeader().setCascadingSectionResizes(False)
        self.tableWidget.horizontalHeader().setDefaultSectionSize(195)
        self.tableWidget.horizontalHeader().setHighlightSections(False)
        self.tableWidget.horizontalHeader().setMinimumSectionSize(45)
        self.tableWidget.horizontalHeader().setSortIndicatorShown(False)
        self.tableWidget.horizontalHeader().setStretchLastSection(False)
        self.tableWidget.verticalHeader().setVisible(True)
        self.tableWidget.verticalHeader().setCascadingSectionResizes(False)
        self.tableWidget.verticalHeader().setDefaultSectionSize(40)
        self.tableWidget.verticalHeader().setHighlightSections(False)
        self.tableWidget.verticalHeader().setMinimumSectionSize(20)
        self.tableWidget.verticalHeader().setSortIndicatorShown(False)
        self.tableWidget.verticalHeader().setStretchLastSection(False)
        self.frame_2 = QtWidgets.QFrame(self.centralwidget)
        self.frame_2.setEnabled(True)
        self.frame_2.setGeometry(QtCore.QRect(220, 0, 511, 579))
        self.frame_2.setStyleSheet("background-color: rgb(86, 179, 240);")
        self.frame_2.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_2.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_2.setObjectName("frame_2")
        self.label_6 = QtWidgets.QLabel(self.frame_2)
        self.label_6.setGeometry(QtCore.QRect(160, 300, 61, 21))
        self.label_6.setStyleSheet("color: rgb(0, 0, 0);\n"
"font: 10pt \"Times New Roman\";")
        self.label_6.setObjectName("label_6")
        self.label_4 = QtWidgets.QLabel(self.frame_2)
        self.label_4.setGeometry(QtCore.QRect(40, 250, 211, 21))
        self.label_4.setStyleSheet("color: rgb(0, 0, 0);\n"
"font: 10pt \"Times New Roman\";")
        self.label_4.setObjectName("label_4")
        self.label_2 = QtWidgets.QLabel(self.frame_2)
        self.label_2.setGeometry(QtCore.QRect(40, 100, 111, 21))
        self.label_2.setStyleSheet("color: rgb(0, 0, 0);\n"
"font: 10pt \"Times New Roman\";")
        self.label_2.setObjectName("label_2")
        self.label_8 = QtWidgets.QLabel(self.frame_2)
        self.label_8.setGeometry(QtCore.QRect(40, 410, 191, 21))
        self.label_8.setStyleSheet("color: rgb(0, 0, 0);\n"
"font: 10pt \"Times New Roman\";")
        self.label_8.setObjectName("label_8")
        self.label_7 = QtWidgets.QLabel(self.frame_2)
        self.label_7.setGeometry(QtCore.QRect(160, 350, 61, 21))
        self.label_7.setStyleSheet("color: rgb(0, 0, 0);\n"
"font: 10pt \"Times New Roman\";")
        self.label_7.setObjectName("label_7")
        self.Tilule = QtWidgets.QLabel(self.frame_2)
        self.Tilule.setGeometry(QtCore.QRect(110, 20, 301, 61))
        self.Tilule.setStyleSheet("font: 14pt \"Times New Roman\";\n"
"color: rgb(0, 0, 0);\n"
"")
        self.Tilule.setObjectName("Tilule")
        self.label_3 = QtWidgets.QLabel(self.frame_2)
        self.label_3.setGeometry(QtCore.QRect(40, 150, 121, 21))
        self.label_3.setStyleSheet("color: rgb(0, 0, 0);\n"
"font: 10pt \"Times New Roman\";")
        self.label_3.setObjectName("label_3")
        self.label_5 = QtWidgets.QLabel(self.frame_2)
        self.label_5.setGeometry(QtCore.QRect(40, 200, 151, 21))
        self.label_5.setStyleSheet("color: rgb(0, 0, 0);\n"
"font: 10pt \"Times New Roman\";")
        self.label_5.setObjectName("label_5")
        self.Button_Calcular = QtWidgets.QPushButton(self.frame_2)
        self.Button_Calcular.setGeometry(QtCore.QRect(250, 470, 211, 31))
        self.Button_Calcular.setStyleSheet("QPushButton{\n"
"background-color: rgb(255, 255, 255);\n"
"color: rgb(0, 0, 0);\n"
"font: 10pt \"Times New Roman\";\n"
"}\n"
"\n"
"QPushButton:hover{\n"
"background-color: rgb(118, 118, 118);\n"
"color: rgb(255, 255, 255);\n"
"font: 10pt \"Times New Roman\";\n"
"}")
        
        self.label_error = QtWidgets.QLabel(self.frame_2)
        self.label_error.setGeometry(QtCore.QRect(51, 530, 410, 31))
        self.label_error.setStyleSheet("color: rgb(0, 0, 0);\n"
"font: 10pt \"Times New Roman\";")
        
        self.label_error.setObjectName("label_error")
        self.Button_Calcular.setObjectName("Button_Calcular")
        self.Button_Calcular.clicked.connect(self.Tomar_datos)
        self.Entry_HP = QtWidgets.QLineEdit(self.frame_2)
        self.Entry_HP.setGeometry(QtCore.QRect(260, 90, 201, 31))
        self.Entry_HP.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.Entry_HP.setObjectName("Entry_HP")
        self.Entry_n = QtWidgets.QLineEdit(self.frame_2)
        self.Entry_n.setGeometry(QtCore.QRect(260, 140, 201, 31))
        self.Entry_n.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.Entry_n.setObjectName("Entry_n")
        self.Entry_Factor = QtWidgets.QLineEdit(self.frame_2)
        self.Entry_Factor.setGeometry(QtCore.QRect(260, 190, 201, 31))
        self.Entry_Factor.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.Entry_Factor.setObjectName("Entry_Factor")
        self.Dim_m = QtWidgets.QLineEdit(self.frame_2)
        self.Dim_m.setGeometry(QtCore.QRect(250, 340, 201, 31))
        self.Dim_m.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.Dim_m.setObjectName("Dim_m")
        self.Tipo_maq = QtWidgets.QLineEdit(self.frame_2)
        self.Tipo_maq.setGeometry(QtCore.QRect(250, 400, 201, 31))
        self.Tipo_maq.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.Tipo_maq.setObjectName("Tipo_maq")
        self.Dim_M = QtWidgets.QLineEdit(self.frame_2)
        self.Dim_M.setGeometry(QtCore.QRect(250, 290, 201, 31))
        self.Dim_M.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.Dim_M.setObjectName("Dim_M")
        self.label = QtWidgets.QLabel(self.frame_2)
        self.label.setGeometry(QtCore.QRect(420, 10, 61, 61))
        self.label.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.label.setText("")
        self.label.setPixmap(QtGui.QPixmap("../../../CEIM 2023/Logo Ing. Mecatrónica UNT.png"))
        self.label.setScaledContents(True)
        self.label.setObjectName("label")
        self.frame_3 = QtWidgets.QFrame(self.centralwidget)
        self.frame_3.setGeometry(QtCore.QRect(220, 0, 511, 579))
        self.frame_3.setStyleSheet("background-color: rgb(86, 179, 240);")
        self.frame_3.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frame_3.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frame_3.setObjectName("frame_3")
        self.label_9 = QtWidgets.QLabel(self.frame_3)
        self.label_9.setGeometry(QtCore.QRect(60, 160, 131, 16))
        self.label_9.setStyleSheet("color: rgb(0, 0, 0);\n"
"font: 10pt \"Times New Roman\";")
        self.label_9.setObjectName("label_9")
        self.label_11 = QtWidgets.QLabel(self.frame_3)
        self.label_11.setGeometry(QtCore.QRect(60, 310, 161, 21))
        self.label_11.setStyleSheet("color: rgb(0, 0, 0);\n"
"font: 10pt \"Times New Roman\";")
        self.label_11.setObjectName("label_11")
        self.Tilule_2 = QtWidgets.QLabel(self.frame_3)
        self.Tilule_2.setGeometry(QtCore.QRect(110, 40, 321, 61))
        self.Tilule_2.setStyleSheet("font: 14pt \"Times New Roman\";\n"
"color: rgb(0,0,0);\n"
"")
        self.Tilule_2.setObjectName("Tilule_2")
        self.label_12 = QtWidgets.QLabel(self.frame_3)
        self.label_12.setGeometry(QtCore.QRect(60, 210, 161, 16))
        self.label_12.setStyleSheet("color: rgb(0, 0, 0);\n"
"font: 10pt \"Times New Roman\";")
        self.label_12.setObjectName("label_12")
        self.label_13 = QtWidgets.QLabel(self.frame_3)
        self.label_13.setGeometry(QtCore.QRect(60, 260, 151, 21))
        self.label_13.setStyleSheet("color: rgb(0, 0, 0);\n"
"font: 10pt \"Times New Roman\";")
        self.label_13.setObjectName("label_13")
        self.LABEL_20 = QtWidgets.QLabel(self.frame_3)
        self.LABEL_20.setGeometry(QtCore.QRect(51, 530, 410, 31))
        self.LABEL_20.setStyleSheet("color: rgb(0, 0, 0);\n"
"font: 10pt \"Times New Roman\";")
        self.Button_Guardar = QtWidgets.QPushButton(self.frame_3)
        self.Button_Guardar.setGeometry(QtCore.QRect(160, 390, 201, 31))
        self.Button_Guardar.setStyleSheet("QPushButton{\n"
"background-color: rgb(255, 255, 255);\n"
"color: rgb(0, 0, 0);\n"
"font: 10pt \"Times New Roman\";\n"
"}\n"
"\n"
"QPushButton:hover{\n"
"background-color: rgb(118, 118, 118);\n"
"color: rgb(255, 255, 255);\n"
"font: 10pt \"Times New Roman\";\n"
"}")
        self.Button_Guardar.setObjectName("Button_Guardar")
        self.Button_Guardar.clicked.connect(self.Guardar)
        self.Long_paso = QtWidgets.QLineEdit(self.frame_3)
        self.Long_paso.setGeometry(QtCore.QRect(250, 150, 201, 31))
        self.Long_paso.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.Long_paso.setObjectName("Long_paso")
        self.Dis_centro = QtWidgets.QLineEdit(self.frame_3)
        self.Dis_centro.setGeometry(QtCore.QRect(250, 200, 201, 31))
        self.Dis_centro.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.Dis_centro.setObjectName("Dis_centro")
        self.Tip_banda = QtWidgets.QLineEdit(self.frame_3)
        self.Tip_banda.setGeometry(QtCore.QRect(250, 250, 201, 31))
        self.Tip_banda.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.Tip_banda.setObjectName("Tip_banda")
        self.Num_banda = QtWidgets.QLineEdit(self.frame_3)
        self.Num_banda.setGeometry(QtCore.QRect(250, 300, 201, 31))
        self.Num_banda.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.Num_banda.setObjectName("Num_banda")
        self.label_17 = QtWidgets.QLabel(self.frame_3)
        self.label_17.setGeometry(QtCore.QRect(430, 20, 61, 61))
        self.label_17.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.label_17.setText("")
        self.label_17.setPixmap(QtGui.QPixmap("../../../CEIM 2023/Logo Ing. Mecatrónica UNT.png"))
        self.label_17.setScaledContents(True)
        self.label_17.setObjectName("label_17")
        MainWindow_principal.setCentralWidget(self.centralwidget)

        self.retranslateUi(MainWindow_principal)
        self.Button_TAB_RES.clicked.connect(self.abrir_Tablas)
        self.Button_Ingresa.clicked.connect(self.frame_2.show)
        self.Button_Ingresa.clicked.connect(self.frame_3.hide)
        self.Button_Ingresa.clicked.connect(self.frame_4.hide)
        self.Button_Resultado.clicked.connect(self.frame_2.hide)
        self.Button_Resultado.clicked.connect(self.frame_3.show)
        self.Button_Resultado.clicked.connect(self.frame_4.hide)
        self.Button_Tip_maq.clicked.connect(self.frame_2.hide)
        self.Button_Tip_maq.clicked.connect(self.frame_3.hide)
        self.Button_Tip_maq.clicked.connect(self.frame_4.show)
        QtCore.QMetaObject.connectSlotsByName(MainWindow_principal)
        MainWindow_principal.setTabOrder(self.Button_Resultado, self.Button_Tip_maq)
        MainWindow_principal.setTabOrder(self.Button_Tip_maq, self.Button_Guardar)
        MainWindow_principal.setTabOrder(self.Button_Guardar, self.Long_paso)
        MainWindow_principal.setTabOrder(self.Long_paso, self.Dis_centro)
        MainWindow_principal.setTabOrder(self.Dis_centro, self.Tip_banda)
        MainWindow_principal.setTabOrder(self.Tip_banda, self.Num_banda)
        MainWindow_principal.setTabOrder(self.Num_banda, self.tableWidget)
        MainWindow_principal.setTabOrder(self.tableWidget, self.Button_Calcular)
        MainWindow_principal.setTabOrder(self.Button_Calcular, self.Entry_HP)
        MainWindow_principal.setTabOrder(self.Entry_HP, self.Entry_n)
        MainWindow_principal.setTabOrder(self.Entry_n, self.Entry_Factor)
        MainWindow_principal.setTabOrder(self.Entry_Factor, self.Dim_m)
        MainWindow_principal.setTabOrder(self.Dim_m, self.Tipo_maq)
        MainWindow_principal.setTabOrder(self.Tipo_maq, self.Dim_M)
        MainWindow_principal.setTabOrder(self.Dim_M, self.Button_Ingresa)
        

    def retranslateUi(self, MainWindow_principal):
        _translate = QtCore.QCoreApplication.translate
        MainWindow_principal.setWindowTitle(_translate("MainWindow_principal", "CALCULADOR DE FAJAS"))
        self.Button_Ingresa.setText(_translate("MainWindow_principal", "INGRESAR DATOS"))
        self.Button_Resultado.setText(_translate("MainWindow_principal", "RESULTADO"))
        self.Button_Tip_maq.setText(_translate("MainWindow_principal", "TIPO DE MAQUINA"))
        self.Button_TAB_RES.setText(_translate("MainWindow_principal", "TABLAS RESTANTES"))
        self.label_18.setText(_translate("MainWindow_principal", "Tabla de tipo de Máquina"))
        self.tableWidget.setToolTip(_translate("MainWindow_principal", "<html><head/><body><p align=\"center\"><br/></p></body></html>"))
        self.tableWidget.setSortingEnabled(False)
        item = self.tableWidget.verticalHeaderItem(0)
        item.setText(_translate("MainWindow_principal", "Uniforme"))
        item = self.tableWidget.verticalHeaderItem(1)
        item.setText(_translate("MainWindow_principal", "Inpacto Ligero"))
        item = self.tableWidget.verticalHeaderItem(2)
        item.setText(_translate("MainWindow_principal", "Inpacto Medio"))
        item = self.tableWidget.verticalHeaderItem(3)
        item.setText(_translate("MainWindow_principal", "Inpacto Pesado"))
        item = self.tableWidget.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow_principal", "Carac. Par de T. Nor."))
        item = self.tableWidget.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow_principal", "Par Tor. alto y no Uni"))
        __sortingEnabled = self.tableWidget.isSortingEnabled()
        self.tableWidget.setSortingEnabled(False)
        item = self.tableWidget.item(0, 0)
        item.setText(_translate("MainWindow_principal", "1.0 a 1.2"))
        item = self.tableWidget.item(1, 0)
        item.setText(_translate("MainWindow_principal", "1.1 a 1.3"))
        item = self.tableWidget.item(1, 1)
        item.setText(_translate("MainWindow_principal", "1.2 a 1.4"))
        item = self.tableWidget.item(2, 0)
        item.setText(_translate("MainWindow_principal", "1.2 a 1.4"))
        item = self.tableWidget.item(2, 1)
        item.setText(_translate("MainWindow_principal", "1.4 a 1.6"))
        item = self.tableWidget.item(3, 0)
        item.setText(_translate("MainWindow_principal", "1.3 a 1.5"))
        item = self.tableWidget.item(3, 1)
        item.setText(_translate("MainWindow_principal", "1.5 a 1.8"))
        self.tableWidget.setSortingEnabled(__sortingEnabled)
        self.label_6.setText(_translate("MainWindow_principal", "Mayor:"))
        self.label_4.setText(_translate("MainWindow_principal", "Diametros de las poleas [pulg]"))
        self.label_2.setText(_translate("MainWindow_principal", "Potencia [HP]"))
        self.label_8.setText(_translate("MainWindow_principal", "Tipo de maquina accionada"))
        self.label_7.setText(_translate("MainWindow_principal", "Menor:"))
        self.Tilule.setText(_translate("MainWindow_principal", "SELECIONADOR DE FAJAS"))
        self.label_3.setText(_translate("MainWindow_principal", "Velocidad [RPM]"))
        self.label_5.setText(_translate("MainWindow_principal", "Factor de Seguridad"))
        self.Button_Calcular.setText(_translate("MainWindow_principal", "CALCULAR"))
        self.label_9.setText(_translate("MainWindow_principal", "Longitud de paso"))
        self.label_11.setText(_translate("MainWindow_principal", "N° de bandas"))
        self.Tilule_2.setToolTip(_translate("MainWindow_principal", "<html><head/><body><p align=\"center\">RESULTADOS OBTENIDOS</p></body></html>"))
        self.Tilule_2.setText(_translate("MainWindow_principal", "RESULTADOS OBTENIDOS"))
        self.label_12.setText(_translate("MainWindow_principal", "Distancia entre centros"))
        self.label_13.setText(_translate("MainWindow_principal", "Tipo de bandas"))
        self.label_error.setText(_translate("MainWindow_principal", "     "))
        self.LABEL_20.setText(_translate("MainWindow_principal", "     "))
        #self.label_error = QtWidgets.QLabel("ESo raro")
        self.Button_Guardar.setText(_translate("MainWindow_principal", "GUARDAR"))

    def Tomar_datos(self):
        self.To_HP  = self.Entry_HP.text()
        self.To_Vel = self.Entry_n.text()
        self.To_FS  = self.Entry_Factor.text()
        self.To_M   = self.Dim_M.text()
        self.To_m   = self.Dim_m.text()
        self.To_TP  = self.Tipo_maq.text()
        if self.To_HP != '' and self.To_Vel != '' and self.To_FS != '' and self.To_M != '' and self.To_m != '' and self.To_TP != '':
                self.To_HP  = int(self.To_HP)
                self.To_Vel = int(self.To_Vel)
                self.To_FS  = float(self.To_FS)
                self.To_M   = float(self.To_M)
                self.To_m   = np.floor(float(self.To_m))
                self.To_TP  = float(self.To_TP)
                print('\nCalculo de fajas \nValores introducidos')
                print('Hp',self.To_HP, ', RPM',self.To_Vel, ', FS',self.To_FS, ', D_M',self.To_M, ', D_m',self.To_m, ', TP',self.To_TP)
                self.calculos_totales()
        else:
                self.label_error.setText(QtCore.QCoreApplication.translate("MainWindow_principal", "NO LLENASTE TODO LOS DATOS :("))

    def calculos_totales(self):
        self.Lp = None
        self.C = None 
        self.N_b = None
        while True:
                # RANGO DE POTENCIA
                if self.To_HP >=15 and self.To_HP <= 100:
                       pass
                else:
                       self.label_error.setText(QtCore.QCoreApplication.translate("MainWindow_principal", "RANGO DE POTENCIA [15 a 100 HP] :("))
                       break

                #RANGO DE TIPO DE MAQUINA
                if self.To_TP>= 1.0 and self.To_TP<=1.8:
                       pass
                else:
                       self.label_error.setText(QtCore.QCoreApplication.translate("MainWindow_principal", "NO EXISTE ESE TIPO DE MAQUINA :("))
                       break

                # Calcular la longitud de la correa
                self.L = self.longitud(self.To_FS, self.To_M, self.To_m)

                # Obtener la longitud de la faja más cercana
                self.long_faja = self.longitud_de_faja(self.L, Tabla_17_10)
                print("La longitud de faja es:", self.long_faja)
                if self.long_faja == None:
                        print("Ingrese nuevos valores")
                        self.label_error.setText(QtCore.QCoreApplication.translate("MainWindow_principal", "NO SE ENCONTRO FAJA PARA ESTAS DIMENCIONES :("))
                        break
                        
                # Calcular la longitud de paso Lp
                self.Lp = self.long_faja + 2.9
                print("La longitud de paso es: ", round(self.Lp, 3))

                # Calcular la distancia entre centros C
                self.C = self.distancia_centros(self.Lp, self.To_M, self.To_m)
                self.C = round(self.C,1)
                print(f'La distancia entre centros C es: {round(self.C, 3)}')

                # Calcular la velocidad de la banda V
                self.V = self.velocidad_banda(self.To_m, self.To_Vel)
                print(f'V: {round(self.V, 3)}')
                if self.V > 5000 :
                        print("NO CONTAMOS CON BANDAS PARA ESTÁ VELOCIDAD")
                        self.label_error.setText(QtCore.QCoreApplication.translate("MainWindow_principal", "NO CONTAMOS CON BANDAS PARA ESTÁ VELOCIDAD :("))
                        break

                # Calcular el factor K1
                self.k1 = self.factor_K1((self.To_M - self.To_m) / self.C)
                if self.k1 is not None:
                        print(f'El valor del factor K1 es: {round(self.k1, 3)}')
                else: 
                        print(f'No se encontró un valor para K1 dentro de la tabla, asegúrese de ingresar los datos correctamente.')

                # Obtener el valor de H_tab
                self.h_tab = self.H_TAB(self.V , self.To_m)
                print(f'H_Tab: {round(self.h_tab, 3)}')

                # Calcular el factor k2
                self.k2 = self.factor_k2(Tabla17_14, self.Lp)

                # Calcular H_a
                self.h_a = self.H_a(self.h_tab, self.k1, self.k2)
                print(f'H_a: {round(self.h_a, 3)}')

                # Calcular la potencia de diseño H_d
                self.h_d = self.potencia_diseño(self.To_HP, self.To_TP, self.To_FS)
                print(f'Potencia de diseño H_d: {self.h_d} HP')

                # Calcular la cantidad de bandas necesarias
                self.N_b = self.cantidad_bandas(self.h_d, self.h_a)
                print(f'Se necesitan {self.N_b} bandas.')
                break
        if self.Lp == None or  self.C == None or self.N_b == None:
               pass
        else:
               self.Mostrar_resultados_f()
        
    def Mostrar_resultados_f(self):
        self.Long_paso.setText(str(self.Lp))
        self.Dis_centro.setText(str(self.C))
        self.Tip_banda.setText("Banda tipo C")
        self.Num_banda.setText(str(self.N_b))

    def Guardar(self):
        try:
            wb = load_workbook("Tabla.xlsx")
        except FileNotFoundError:
            wb = Workbook()

        ws = wb.active

        ws.cell(row=1, column=1, value="Longitud de paso")
        ws.cell(row=1, column=2, value="Distancia entre centros")
        ws.cell(row=1, column=3, value="Tipo de banda")
        ws.cell(row=1, column=4, value="Numero de bandas")

        # Obtener la última fila con datos
        last_row = ws.max_row + 1

        # Datos a guardar en la tabla
        datos = [
            [self.Lp, self.C, "Banda tipo C", self.N_b],
        ]

        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 22
        ws.column_dimensions['D'].width = 22
        # Guardar los datos en la tabla
        for row_data in datos:
            for col_idx, col_data in enumerate(row_data, start=1):
                cell = ws.cell(row=last_row, column=col_idx, value=str(col_data))
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = cell.border + Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"))

            last_row += 1

        # Guardar el archivo de Excel
        wb.save("Tabla.xlsx")
        self.LABEL_20.setText(QtCore.QCoreApplication.translate("MainWindow_principal", "MINIMO MI 20 :) "))

    def abrir_Tablas(self):
        self.ventana = QtWidgets.QMainWindow()
        self.Ui = ImageGallery()
        self.Ui.show()

    def longitud(self,n_d, D, d):
        y = n_d * np.pi * (D + d)
        return y

        # Función para obtener la longitud de la faja más cercana a un valor dado
    def longitud_de_faja(self, dato, tabla):
        long_faja = None
        
        if dato in tabla:  # Si el dato está presente en la tabla, se selecciona ese número
                long_faja = dato
        else:
                mayores = [num for num in tabla if num > dato] 
                if mayores:  # Si hay números mayores
                        long_faja = min(mayores)  # Selecciona el número más cercano
        
        return long_faja

    # Función para calcular la distancia entre los centros de los diámetros de la correa
    def distancia_centros(self, Lp, D, d):
        C = 0.25 * ((Lp - (np.pi/2) * (D + d)) + np.sqrt(np.power(Lp - (np.pi/2) * (D + d), 2) - 2 * np.power(D - d, 2)))
        return C

    # Función para verificar si el valor de la distancia entre centros está en el rango correcto
    def ajustar_rango(C, D, d):
        if D < C and C < 3 * (D + d):
                print("El valor de C está en el rango correcto")
        else:
                print("El valor de C NO está en el rango correcto")

     # Función para calcular la velocidad de la banda
    def velocidad_banda(self, d, n):
        return (np.pi * d * n) / 12

# Función para obtener el factor K1 de la tabla Tabla_17_13
    def factor_K1(self, valor):
        for i in range(len(Tabla_17_13)):
                if Tabla_17_13[i]["(D-d)\/C"] == valor:
                        return Tabla_17_13[i]["K1_VV"]
                elif Tabla_17_13[i]["(D-d)\/C"] > valor:
                        sup = i
                        inf = i - 1
                        break
        if sup is not None and inf is not None:
                k1 = self.interpolacion(Tabla_17_13[inf]["(D-d)\/C"], Tabla_17_13[inf]["K1_VV"], Tabla_17_13[sup]["(D-d)\/C"], Tabla_17_13[sup]["K1_VV"], valor)
                return k1
        else:
                k1 = None

# Función para obtener el factor k2 de la tabla Tabla17_14
    def factor_k2(self, tabla17_14, longitud):
        for i in range(len(tabla17_14)):
                if tabla17_14[i]["L_min"] <= longitud <= tabla17_14[i]["L_max"]:
                        return tabla17_14[i]["Factor"]

        # Si no se encuentra en ningún intervalo, realizar interpolación
        l_mins = [t["L_min"] for t in tabla17_14]
        l_maxs = [t["L_max"] for t in tabla17_14]

        # Encontrar los valores más cercanos de L_max y L_min
        l_max_anterior = max(filter(lambda x: x <= longitud, l_maxs))
        l_min_superior = min(filter(lambda x: x >= longitud, l_mins))

        # Encontrar los factores correspondientes
        factor_lmax_anterior = [t["Factor"] for t in tabla17_14 if t["L_max"] == l_max_anterior][0]
        factor_lmin_superior = [t["Factor"] for t in tabla17_14 if t["L_min"] == l_min_superior][0]

        # Realizar interpolación
        factor_interpolado = factor_lmax_anterior + (factor_lmin_superior - factor_lmax_anterior) * \
                                (longitud - l_max_anterior) / (l_min_superior - l_max_anterior)

        return factor_interpolado

# Función para obtener el valor de H_tab de la tabla Tabla_17_12_C
    def H_TAB(self, V, d):
        # Si el valor de d es mayor a 12, se toma como 12
        if d > 12:
                d = 12

        if V > 5000:
                return None
        for i in range(5):
                velocidad = Tabla_17_12_C[i][0]

                if velocidad == V:
                        for j in range(8):
                                j += 1 
                                if d == Tabla_17_12_C[i][j]["Diámetro"]:
                                        H_tab = Tabla_17_12_C[i][j]["H_tab"]
                                        return H_tab
                        return None
                
                elif velocidad > V:
                        sup = i
                        inf = i - 1
                        break

        if sup is not None and inf is not None:
                for k in range(7):
                        k += 1
                        if d == Tabla_17_12_C[inf][k]["Diámetro"]:
                                h_inf = Tabla_17_12_C[inf][k]["H_tab"]

                        if d == Tabla_17_12_C[sup][k]["Diámetro"]:
                                h_sup = Tabla_17_12_C[sup][k]["H_tab"]
                print(h_inf, h_sup)
                if h_inf is not None and h_sup is not None:
                        H_tab = self.interpolacion(Tabla_17_12_C[inf][0], h_inf, Tabla_17_12_C[sup][0], h_sup, V)
                        return H_tab

        print("No se encontró el valor de V en la tabla.")
        return None

# Función para calcular H_a multiplicando H_tab, k1 y k2
    def H_a(self, H_tab, k1, k2):
        return H_tab * k1 * k2

# Función para obtener el factor de servicio Ks según el tipo de máquina y par de torsión

    def factor_servicio(tipo):
        while True:
                try:
                        maquina = int(input("\nElija maquinaria impulsada:\n1. Uniforme.\n2. Impacto ligero.\n3. Impacto medio.\n4. Impacto pesado\nElección: "))
                        if maquina != 1 and maquina != 2 and maquina != 3 and maquina != 4:
                                print("Ingrese una opción válida")
                        else: 
                                break
                except (ValueError):
                        print("Ingrese una opción válida")

        if tipo == 1:
                if maquina == 1:
                        Ks = Tabla_17_15[0][0]["Ks"]
                elif maquina == 2:
                        Ks = Tabla_17_15[0][1]["Ks"]
                elif maquina == 3:
                        Ks = Tabla_17_15[0][2]["Ks"]
                elif maquina == 4:
                        Ks = Tabla_17_15[0][3]["Ks"]
        elif tipo == 2:
                if maquina == 1:
                        Ks = Tabla_17_15[1][0]["Ks"]
                elif maquina == 2:
                        Ks = Tabla_17_15[1][1]["Ks"]
                elif maquina == 3:
                        Ks = Tabla_17_15[1][2]["Ks"]
                elif maquina == 4:
                        Ks = Tabla_17_15[1][3]["Ks"]
        return Ks

# Función para calcular la potencia de diseño multiplicando H_nom, Ks y n_d
    def potencia_diseño(self, H_nom, Ks, n_d):
        H_d = H_nom * Ks * n_d
        return H_d

# Función para calcular la cantidad de bandas necesarias para H_d y H_a
    def cantidad_bandas(self, H_d, H_a):
        num_bandas = H_d / H_a
        if num_bandas.is_integer():
                return int(num_bandas)
        else:
                return np.ceil(num_bandas)

# Función para realizar la interpolación lineal entre dos puntos (x1, y1) y (x2, y2) para un valor x dado
    def interpolacion(self,x1, y1, x2, y2, x):
        y = y1 + (x - x1) * (y2 - y1) / (x2 - x1)
        return y
        



if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow_principal = QtWidgets.QMainWindow()
    ui = Ui_MainWindow_principal()
    ui.setupUi(MainWindow_principal)
    MainWindow_principal.show()
    sys.exit(app.exec_())

